"""Build kandy_core_vocab.json from kandy_core_vocab.xlsx."""
import json
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
XLSX_PATH = ROOT / "kandy_core_vocab.xlsx"
JSON_PATH = ROOT / "kandy_core_vocab.json"


def read_rows():
    wb = load_workbook(XLSX_PATH, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        keyword, word, kanji, hiragana, _blocked, distractors = row[:6]
        rows.append({
            "keyword": keyword,
            "word": word,
            "kanji_raw": kanji,
            "hiragana_raw": hiragana,
            "distractors_raw": distractors,
        })
    return rows


def split_csv(value):
    if not value:
        return []
    return [p.strip() for p in value.split(",") if p.strip()]


def build():
    rows = read_rows()

    # Homographic resolution: first-row occurrence wins.
    # Map kanji-form (word column) -> keyword of first occurrence.
    word_to_keyword = {}
    for row in rows:
        if row["word"] not in word_to_keyword:
            word_to_keyword[row["word"]] = row["keyword"]

    # Build entries and resolve distractors.
    seen_keywords = set()
    entries = []
    unresolved = []
    for row in rows:
        kw = row["keyword"]
        if kw in seen_keywords:
            print(f"ERROR: duplicate keyword '{kw}'", file=sys.stderr)
            sys.exit(1)
        seen_keywords.add(kw)

        kanji_tokens = split_csv(row["kanji_raw"])
        hiragana_tokens = split_csv(row["hiragana_raw"])
        distractor_tokens = split_csv(row["distractors_raw"])

        resolved = []
        for d_kanji in distractor_tokens:
            target_kw = word_to_keyword.get(d_kanji)
            if target_kw is None:
                unresolved.append((kw, d_kanji))
                continue
            resolved.append(target_kw)

        entries.append({
            "keyword": kw,
            "word": row["word"],
            "kanji": kanji_tokens,
            "hiragana": hiragana_tokens,
            "distractors": resolved,
        })

    if unresolved:
        print(f"ERROR: {len(unresolved)} unresolved distractors:", file=sys.stderr)
        for kw, d in unresolved[:20]:
            print(f"  in '{kw}': '{d}'", file=sys.stderr)
        sys.exit(1)

    return {"version": 1, "words": entries}


def main():
    payload = build()
    with JSON_PATH.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    print(f"Wrote {len(payload['words'])} entries to {JSON_PATH.name}")


if __name__ == "__main__":
    main()
